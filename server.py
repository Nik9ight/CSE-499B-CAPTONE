"""
Segmentation API for RadVision

Two route styles — use whichever matches your frontend config:

  POST /segment/<image_type>   e.g. /segment/chest_xray  (type comes from URL)
  POST /segment                fallback; defaults to chest_xray

Request body:  { "image_base64": "<base64 string>" }
Response:      { "mask_base64": "<base64 binary PNG>" }

White pixels (brightness > 128) = segmented region
Black pixels = background

To add a new modality, edit model_registry.py.
"""

import base64
import io
import os
from flask import Flask, request, jsonify
from flask_cors import CORS
from PIL import Image
from explain import explain
from test_model import test_model
from reportGen import reportGen
from annotation import get_medgemma_description
from transformers import AutoProcessor, AutoModelForImageTextToText
from huggingface_hub import login
import torch
# Load .env if present (pip install python-dotenv, or set vars in your shell)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv optional — set env vars in your shell instead
MODEL_ID = "google/medgemma-1.5-4b-it"
isON = False
model= "None"
processor = "None"
def load_model():
    """Load model and processor once."""
    hf_token = os.environ.get("HF_TOKEN")
    if not hf_token:
        raise RuntimeError("HF_TOKEN is not set. Add it to .env or environment variables.")

    login(token=hf_token)

    dtype = torch.bfloat16 if torch.cuda.is_available() else torch.float32
    device_map = "cuda" if torch.cuda.is_available() else "cpu"

    model = AutoModelForImageTextToText.from_pretrained(
        MODEL_ID,
        dtype=dtype,
        device_map=device_map,
    )
    processor = AutoProcessor.from_pretrained(MODEL_ID)
    return model, processor
model, processor = load_model()

import base64
import numpy as np
from PIL import Image
from io import BytesIO


def overlay_mask(image_b64: str, mask_b64: str, color=(0, 200, 100), opacity=0.3) -> str:
    """
    Overlays a green highlight on an image where the mask is white.

    Args:
        image_b64: Base64-encoded original image (224x224).
        mask_b64:  Base64-encoded black/white mask (224x224).
        color:     RGB tuple for the overlay color. Default green.
        opacity:   Opacity of the overlay (0.0 - 1.0). Default 0.6.

    Returns:
        Base64-encoded overlay image (PNG).
    """
    img = Image.open(BytesIO(base64.b64decode(image_b64))).convert("RGBA")
    mask = Image.open(BytesIO(base64.b64decode(mask_b64))).convert("L")

    img = img.resize((224, 224))
    mask = mask.resize((224, 224))

    mask_arr = np.array(mask)
    overlay = np.array(img).copy()

    white = mask_arr > 127

    for c, val in enumerate(color):
        overlay[white, c] = (
            overlay[white, c] * (1 - opacity) + val * opacity
        ).astype(np.uint8)

    result = Image.fromarray(overlay, "RGBA")

    buf = BytesIO()
    result.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")

HOST  = os.environ.get("FLASK_HOST",  "0.0.0.0")
PORT  = int(os.environ.get("FLASK_PORT",  "5000"))
DEBUG = os.environ.get("FLASK_DEBUG", "false").lower() == "true"

app = Flask(__name__)
CORS(app)  # Required so the browser can call this from the HTML file

def get_image_description(image_filename, modality):
    """Return the Description for image_filename from models/{modality}.xlsx."""
    xlsx_path = os.path.join("models", f"{modality}.xlsx")
    if not os.path.exists(xlsx_path):
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            img_col = headers.index("Image")
            desc_col = headers.index("Description")
        except ValueError:
            return None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[img_col] == image_filename:
                return row[desc_col]
    except Exception as e:
        print(f"Error reading {xlsx_path}: {e}")
    return None


def segmentation_presets(image_filename):
    if not image_filename:
        return None
    maskname = "mask_" + image_filename
    mask_path = os.path.join("./models/assets/", maskname)
    if os.path.exists(mask_path):
        return Image.open(mask_path).convert("RGB")

@app.route("/hello", methods=["GET"])
def hello():
    return "Hello from RadVision backend!"


@app.route("/segment", methods=["POST"])
def segment():
    """Route that reads modality from query param, e.g. ?modality=chest_xray"""
    modality = request.args.get("modality", "chest_xray")  # defaults to chest_xray if not provided
    image_filename = request.args.get("image_filename", "uploaded_image.png")  # optional filename for logging
    print(f"Received segmentation request for modality '{modality}' with image filename '{image_filename}'")
    if image_filename and image_filename.startswith("preset_"):
        print("its a preset image!")
        preset_mask =  segmentation_presets(image_filename)
        if preset_mask:
            buf = io.BytesIO()
            preset_mask.save(buf, format="PNG")
            mask_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
            return jsonify({"mask_base64": mask_b64})
    if image_filename and isON:
        maskname = "mask_" + image_filename 
        mask_path = os.path.join("models", modality, maskname)
        print(f"Looking for existing mask at '{mask_path}'...")
        if os.path.exists(mask_path):
            print(f"Found existing mask at '{mask_path}'. Returning it without inference.")
            mask_img = Image.open(mask_path).convert("RGB")
            buf = io.BytesIO()
            mask_img.save(buf, format="PNG")
            mask_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
            return jsonify({"mask_base64": mask_b64})
    data = request.get_json(force=True)
    if not data or "image_base64" not in data:
        return jsonify({"error": "Missing image_base64 field"}), 400
    try:
        img_bytes = base64.b64decode(data["image_base64"])
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    except Exception as e:
        return jsonify({"error": f"Could not decode image: {e}"}), 400
    # Read from the Excel file
    maskname = "mask_" + image_filename 
    mask_path = os.path.join("models", modality, maskname)
    if os.path.exists(mask_path):
        text = get_image_description(maskname, modality)
    else:
        print(f"No existing mask found for '{maskname}' in modality '{modality}'. Running model inference.")
        text = get_medgemma_description(img, model, processor)
    mask_img = test_model(img, text=text, modality=modality)

    buf = io.BytesIO()
    mask_img.save(buf, format="PNG")
    mask_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return jsonify({"mask_base64": mask_b64})



@app.route("/report", methods=["POST"])
def generate_report():
    data = request.get_json(force=True)
    if not data or "image_base64" not in data:
        return jsonify({"error": "Missing image_base64 field"}), 400
    try:
        img_bytes = base64.b64decode(data["image_base64"])
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    except Exception as e:
        return jsonify({"error": f"Could not decode image: {e}"}), 400
    image_filename = request.args.get("image_filename", "uploaded_image.png")  # optional filename for logging
    modality = request.args.get("modality", "chest_xray")  # defaults to chest_xray if not provided
    if image_filename and image_filename.startswith("preset_"):
        print("its a preset image!")
        # Look for a matching .txt report file in ./models/assets/
        base = os.path.splitext(image_filename)[0]
        candidate_paths = [
            os.path.join("./models/assets", base + ".txt"),
            os.path.join("./models/assets", image_filename + ".txt"),
            os.path.join("./models/assets", image_filename),
        ]
        report = None
        for p in candidate_paths:
            if os.path.exists(p):
                try:
                    with open(p, "r", encoding="utf-8") as f:
                        report = f.read()
                except Exception as e:
                    report = f"Could not read preset report file: {e}"
                break
        if report is None:
            report = "Preset report not found."
        return jsonify({"report": report})
    maskname = "mask_" + image_filename 
    mask_path = os.path.join("models", modality, maskname)
    if os.path.exists(mask_path):
        text = get_image_description(maskname, modality)
    else:
        print(f"No existing mask found for '{maskname}' in modality '{modality}'. Running model inference.")
        text = get_medgemma_description(img, model, processor)
    mask_img = test_model(img, text=text, modality=modality)
    buf = io.BytesIO()
    mask_img.save(buf, format="PNG")
    mask_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    overlay = overlay_mask(data["image_base64"], mask_b64)
    report = reportGen(overlay, text, model, processor)
    return jsonify({"report": report})

@app.route("/explain", methods=["POST"])
def explainSentence():
    data = request.get_json(force=True)
    sentence = data.get("sentenceText", "")
    response = explain(sentence, model, processor)
    return jsonify({"explanation": response})

if __name__ == "__main__":
    app.run(host=HOST, port=PORT, debug=DEBUG)
